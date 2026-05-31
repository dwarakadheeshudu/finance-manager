import os
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy.orm import Session
from backend.app.models import Expense, Budget, User, SavingsGoal
from backend.app.services.ai_service import calculate_financial_health_score, generate_savings_recommendations

# Ensure directories exist
REPORTS_DIR = "backend_static_reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_pdf_report(db: Session, user: User, start_date: datetime, end_date: datetime) -> str:
    """
    Generates a professional financial PDF report and saves it to a static reports folder.
    """
    filename = f"financial_report_{user.id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.pdf"
    file_path = os.path.join(REPORTS_DIR, filename)

    # Fetch expenses and data
    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= start_date,
        Expense.date <= end_date
    ).all()

    total_spent = sum(e.amount for e in expenses)
    remaining_balance = user.monthly_salary - total_spent
    
    score, grade, feedback = calculate_financial_health_score(db, user.id)
    recommendations = generate_savings_recommendations(db, user.id)

    doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for Premium Appearance
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0F172A'), # Slate 900
        alignment=0, # Left
        spaceAfter=15
    )

    h2_style = ParagraphStyle(
        'Heading2_Custom',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1E293B'), # Slate 800
        spaceBefore=12,
        spaceAfter=6
    )

    body_style = ParagraphStyle(
        'Body_Custom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#475569'), # Slate 600
        spaceAfter=4
    )

    bold_body_style = ParagraphStyle(
        'BoldBody_Custom',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=4
    )

    # Document Header
    story.append(Paragraph("Smart Personal Finance Manager With AI", title_style))
    story.append(Paragraph(f"Monthly Summary Report | Generated: {datetime.utcnow().strftime('%B %Y')}", body_style))
    story.append(Spacer(1, 10))
    
    # Overview Table
    overview_data = [
        [Paragraph("Metric", bold_body_style), Paragraph("Value (₹)", bold_body_style)],
        [Paragraph("Monthly Income / Salary", body_style), Paragraph(f"{user.monthly_salary:,.2f}", body_style)],
        [Paragraph("Monthly Savings Goal", body_style), Paragraph(f"{user.savings_goal:,.2f}", body_style)],
        [Paragraph("Total Expenses Logged", body_style), Paragraph(f"{total_spent:,.2f}", body_style)],
        [Paragraph("Remaining Balance", body_style), Paragraph(f"{remaining_balance:,.2f}", body_style)],
    ]
    
    t_overview = Table(overview_data, colWidths=[200, 200])
    t_overview.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#0F172A')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(Paragraph("Financial Overview", h2_style))
    story.append(t_overview)
    story.append(Spacer(1, 15))

    # Health Score Card
    health_data = [
        [Paragraph("Financial Health Score", bold_body_style), Paragraph(f"{score} / 100 ({grade})", bold_body_style)],
        [Paragraph("Feedback & Diagnostics", body_style), Paragraph(feedback.replace(' | ', '\n'), body_style)]
    ]
    t_health = Table(health_data, colWidths=[150, 300])
    t_health.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ECFDF5') if score >= 70 else colors.HexColor('#FEF2F2')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1FAE5') if score >= 70 else colors.HexColor('#FEE2E2')),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(Paragraph("AI Diagnostics & Health Status", h2_style))
    story.append(t_health)
    story.append(Spacer(1, 15))

    # Recommendations
    story.append(Paragraph("AI Saving Recommendations", h2_style))
    for rec in recommendations:
        story.append(Paragraph(f"• {rec}", body_style))
    story.append(Spacer(1, 15))

    # Expense breakdown by category
    cat_spent = {}
    for e in expenses:
        cat_spent[e.category] = cat_spent.get(e.category, 0.0) + e.amount
        
    breakdown_data = [[Paragraph("Category", bold_body_style), Paragraph("Spent (₹)", bold_body_style), Paragraph("Percentage", bold_body_style)]]
    for cat, amt in cat_spent.items():
        pct = (amt / total_spent * 100) if total_spent > 0 else 0
        breakdown_data.append([
            Paragraph(cat, body_style),
            Paragraph(f"{amt:,.2f}", body_style),
            Paragraph(f"{pct:.1f}%", body_style)
        ])
        
    t_breakdown = Table(breakdown_data, colWidths=[150, 150, 150])
    t_breakdown.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    story.append(Paragraph("Category Spending Breakdown", h2_style))
    story.append(t_breakdown)

    doc.build(story)
    return file_path

def generate_excel_report(db: Session, user: User, start_date: datetime, end_date: datetime) -> str:
    """
    Generates a styled Excel sheet of expense transaction logs.
    """
    filename = f"financial_report_{user.id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = os.path.join(REPORTS_DIR, filename)

    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= start_date,
        Expense.date <= end_date
    ).order_by(Expense.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Log"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid') # Slate 900
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0F172A')
    normal_font = Font(name='Segoe UI', size=10)
    total_font = Font(name='Segoe UI', size=11, bold=True)
    
    border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title Block
    ws.merge_cells("A1:E1")
    ws["A1"] = "Smart Personal Finance Manager - Expense Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws["A2"] = f"User: {user.full_name} ({user.email})"
    ws["A2"].font = Font(name='Segoe UI', size=10, italic=True)
    
    ws["A3"] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws["A3"].font = Font(name='Segoe UI', size=10, italic=True)
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 20

    # Headers
    headers = ["Date", "Description / Input Text", "Category", "Amount (₹)", "AI Confidence (%)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows
    row_num = 6
    for e in expenses:
        ws.cell(row=row_num, column=1, value=e.date.strftime("%Y-%m-%d %H:%M")).font = normal_font
        ws.cell(row=row_num, column=2, value=e.expense_text or "Manual Log").font = normal_font
        ws.cell(row=row_num, column=3, value=e.category).font = normal_font
        
        amt_cell = ws.cell(row=row_num, column=4, value=e.amount)
        amt_cell.font = normal_font
        amt_cell.number_format = '₹#,##0.00'
        
        conf_cell = ws.cell(row=row_num, column=5, value=e.confidence / 100.0)
        conf_cell.font = normal_font
        conf_cell.number_format = '0.0%'
        
        for col_num in range(1, 6):
            c = ws.cell(row=row_num, column=col_num)
            c.border = thin_border
            if col_num in [1, 3, 5]:
                c.alignment = Alignment(horizontal="center")
            elif col_num == 4:
                c.alignment = Alignment(horizontal="right")
                
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # Totals Row
    ws.cell(row=row_num, column=2, value="Total Spending").font = total_font
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="right")
    
    total_cell = ws.cell(row=row_num, column=4, value=f"=SUM(D6:D{row_num-1})")
    total_cell.font = total_font
    total_cell.number_format = '₹#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    
    # Border for totals
    double_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    ws.cell(row=row_num, column=2).border = double_bottom
    ws.cell(row=row_num, column=4).border = double_bottom

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)
    return file_path
